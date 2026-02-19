#!/usr/bin/env python3
"""
PHG-France — build.py
Génère data.json à partir de l'Excel + scanne les photos.
La structure du JSON est identique au const DATA du HTML existant.

Usage:
    python build.py
    python build.py --verify   (compare avec le HTML existant)
"""

import json
import os
import sys
import re
import unicodedata
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl non installé. Lance : py -m pip install openpyxl")
    sys.exit(1)


# ============================================================
# CONFIG
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(SCRIPT_DIR, "excel")
PHOTOS_DIR = os.path.join(SCRIPT_DIR, "photos")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "data.json")

# Onglets structurels fixes (peuvent être masqués)
FIXED_TABS = {
    "GRANITS",
    "Poids",
    "Tarif TFranco",
    "Zone.TFranco",
    "LISTES",
}

# Onglets ignorés explicitement
IGNORED_TABS = {
    "Mode opératoire",
    "CALCUL PRIX ACHAT €",
    "CALCUL PRIX FAMILLE €",
    "(Semelles.Monument.€HT)",
}

# Pattern pour détecter les onglets produits
PRODUCT_TAB_PATTERN = r".+\.PrixAdh\.€HT$"


# ============================================================
# UTILITAIRES
# ============================================================
def find_excel():
    """Trouve le fichier Excel dans le dossier excel/."""
    if not os.path.isdir(EXCEL_DIR):
        print(f"❌ Dossier '{EXCEL_DIR}' introuvable.")
        sys.exit(1)
    files = [f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx") and not f.startswith("~")]
    if not files:
        print(f"❌ Aucun fichier .xlsx dans '{EXCEL_DIR}'.")
        sys.exit(1)
    if len(files) > 1:
        print(f"⚠️  Plusieurs .xlsx trouvés, utilisation de : {files[0]}")
    return os.path.join(EXCEL_DIR, files[0])


def clean_number(val, decimals=2):
    """Retourne int si le nombre est entier, float sinon. Comme dans le HTML."""
    if val is None:
        return None
    f = round(float(val), decimals)
    if f == int(f):
        return int(f)
    return f


def cell_val(cell):
    """Retourne la valeur d'une cellule, None si vide."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def normalize_ref(ref):
    """Normalise une référence : 'PHGA - CL - A' → 'PHGA-CL-A'."""
    if not ref:
        return ""
    return re.sub(r"\s*-\s*", "-", str(ref).strip())


def normalize_granit_name(code, nom):
    """Normalise le nom d'un granit pour le fichier photo.
    Ex: (31, "Feuille d'automne chinois") → '31-feuille-automne-chinois'
    """
    if not nom:
        return str(code)
    # Minuscule
    name = nom.lower()
    # Supprimer accents
    name = unicodedata.normalize("NFD", name)
    name = "".join(c for c in name if unicodedata.category(c) != "Mn")
    # Supprimer les contractions françaises (d', l', n', qu', etc.)
    name = re.sub(r"\b[dlnqsj][''`]", "", name)
    # Supprimer apostrophes restantes
    name = re.sub(r"[''`]", "", name)
    # Remplacer espaces et / par tirets
    name = re.sub(r"[\s/\\]+", "-", name)
    # Supprimer tout ce qui n'est pas alphanumérique ou tiret
    name = re.sub(r"[^a-z0-9-]", "", name)
    # Nettoyer tirets multiples
    name = re.sub(r"-+", "-", name).strip("-")
    return f"{code}-{name}"


def find_photo(directory, base_name):
    """Cherche une photo (jpg ou png) dans un dossier.
    Retourne le chemin relatif ou None.
    """
    if not os.path.isdir(directory):
        return None
    for ext in [".jpg", ".jpeg", ".png"]:
        filename = base_name + ext
        filepath = os.path.join(directory, filename)
        if os.path.isfile(filepath):
            # Retourner le chemin relatif depuis la racine du projet
            return os.path.relpath(filepath, SCRIPT_DIR).replace("\\", "/")
    return None


# ============================================================
# LECTURE EXCEL — ONGLETS STRUCTURELS
# ============================================================
def read_granits(wb):
    """Lit l'onglet GRANITS → liste de {code, nom, origine}."""
    ws = wb["GRANITS"]
    granits = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        code = cell_val(row[0])  # A: Code granit
        nom = cell_val(row[1])   # B: Granit
        origine = cell_val(row[2])  # C: Origine
        if code is not None and nom:
            # Chercher la photo
            photo_base = normalize_granit_name(code, nom)
            photo = find_photo(os.path.join(PHOTOS_DIR, "granits"), photo_base)
            entry = {
                "code": int(code) if isinstance(code, (int, float)) else code,
                "nom": str(nom),
                "origine": str(origine).strip() if origine else "",
            }
            if photo:
                entry["photo"] = photo
            granits.append(entry)
    return granits


def read_poids(wb):
    """Lit l'onglet Poids → dict {référence: poids_en_tonnes}."""
    ws = wb["Poids"]
    poids = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        ref = cell_val(row[1])   # B: Référence
        val = cell_val(row[2])   # C: Poids en T
        if ref and val is not None:
            poids[str(ref)] = clean_number(val, 3) if val else 0
    return poids


def read_zones_transport(wb):
    """Lit Zone.TFranco → dict {code_dept: nom_zone}."""
    ws = wb["Zone.TFranco"]
    zones = {}
    # Headers : Zone 1, Zone 2, ..., Zone 6 en colonnes A-F
    for row in ws.iter_rows(min_row=2, values_only=False):
        for col_idx in range(6):  # Colonnes A à F
            dept = cell_val(row[col_idx])
            if dept is not None:
                dept_str = str(int(dept)) if isinstance(dept, (int, float)) else str(dept)
                # Formater sur 2 chiffres pour les deps < 10
                if dept_str.isdigit() and len(dept_str) == 1:
                    dept_str = "0" + dept_str
                zone_name = f"Zone {col_idx + 1}"
                zones[dept_str] = zone_name
    return zones


def read_tarifs_transport(wb):
    """Lit Tarif TFranco → liste de {zone, 0_3T, 3_5T, ...}."""
    ws = wb["Tarif TFranco"]
    tarifs = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        zone = cell_val(row[0])  # A: Zone
        if zone and str(zone).startswith("Zone"):
            entry = {
                "zone": str(zone),
                "0_3T": int(row[1].value) if row[1].value else 0,
                "3_5T": int(row[2].value) if row[2].value else 0,
                "5_8T": int(row[3].value) if row[3].value else 0,
                "8_10T": int(row[4].value) if row[4].value else 0,
                "10_15T": int(row[5].value) if row[5].value else 0,
                "minimum": int(row[6].value) if row[6].value else 0,
            }
            tarifs.append(entry)
    return tarifs


def read_listes(wb):
    """Lit l'onglet LISTES → types, lignes_monument, lignes_accessoire, departements."""
    ws = wb["LISTES"]
    types = set()
    lignes_monument = []
    lignes_accessoire = []
    departements = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        dept = cell_val(row[0])   # A: Département
        zone = cell_val(row[1])   # B: Zone
        typ = cell_val(row[2])    # C: Types
        lm = cell_val(row[3])     # D: Lignes_Monument
        la = cell_val(row[4])     # E: Lignes_Accessoire

        if typ:
            types.add(str(typ))
        if lm and str(lm) not in lignes_monument:
            lignes_monument.append(str(lm))
        if la and str(la) not in lignes_accessoire:
            lignes_accessoire.append(str(la))
        if dept is not None and zone:
            dept_str = str(int(dept)) if isinstance(dept, (int, float)) else str(dept)
            if dept_str.isdigit() and len(dept_str) == 1:
                dept_str = "0" + dept_str
            # Éviter les doublons
            if not any(d["departement"] == dept_str for d in departements):
                departements.append({
                    "departement": dept_str,
                    "zone": str(zone)
                })

    # Ordre standard des types
    type_order = ["Monument", "Semelle", "Accessoire", "Urne", "Gravure", "Litho"]
    types_list = [t for t in type_order if t in types]
    # Ajouter les types détectés mais pas dans l'ordre standard
    for t in sorted(types):
        if t not in types_list:
            types_list.append(t)

    return types_list, lignes_monument, lignes_accessoire, departements


# ============================================================
# LECTURE EXCEL — ONGLETS PRODUITS (auto-détection)
# ============================================================
def detect_product_tabs(wb):
    """Détecte les onglets produits visibles matchant *.PrixAdh.€HT."""
    tabs = []
    for name in wb.sheetnames:
        if name in IGNORED_TABS:
            continue
        if name in FIXED_TABS:
            continue
        if re.match(PRODUCT_TAB_PATTERN, name):
            ws = wb[name]
            # Vérifier que l'onglet n'est pas masqué entre parenthèses
            if name.startswith("(") and name.endswith(")"):
                continue
            # Vérifier sheet_state si disponible
            if hasattr(ws, 'sheet_state') and ws.sheet_state != 'visible':
                continue
            tabs.append(name)
    return tabs


def extract_product_type(tab_name):
    """Extrait le type de produit du nom d'onglet.
    'Monument.PrixAdh.€HT' → 'Monument'
    'Accessoires.PrixAdh.€HT' → 'Accessoire'
    """
    base = tab_name.split(".")[0]
    # Normaliser le pluriel
    if base.endswith("s") and base != "Poids":
        base = base[:-1]
    return base


def read_monuments(wb, tab_name):
    """Lit un onglet monuments → liste d'objets."""
    ws = wb[tab_name]
    items = []
    photos_dir = os.path.join(PHOTOS_DIR, "monuments")

    for row in ws.iter_rows(min_row=2, values_only=False):
        ligne = cell_val(row[0])      # A: Ligne
        ref = cell_val(row[1])        # B: Référence
        origine = cell_val(row[2])    # C: I/C
        code_g = cell_val(row[3])     # D: Code granit
        granit = cell_val(row[4])     # E: Granit
        prix = cell_val(row[5])       # F: Prix HT
        sem130 = cell_val(row[6])     # G: Avec semelle 130x230
        sem140 = cell_val(row[7])     # H: Avec semelle 140x240
        sem150 = cell_val(row[8])     # I: Avec semelle 150x250

        if ref and prix is not None:
            entry = {
                "ligne": str(ligne) if ligne else "",
                "reference": str(ref),
                "origine": str(origine) if origine else "",
                "code_granit": int(code_g) if code_g else 0,
                "granit": str(granit) if granit else "",
                "prix_ht": clean_number(prix) if prix else 0,
                "avec_semelle_130x230": clean_number(sem130),
                "avec_semelle_140x240": clean_number(sem140),
                "avec_semelle_150x250": clean_number(sem150),
            }
            # Photo (une par référence, pas par granit)
            photo = find_photo(photos_dir, normalize_ref(ref))
            if photo:
                entry["photo"] = photo
            items.append(entry)
    return items


def read_semelles(wb, tab_name):
    """Lit un onglet semelles → liste d'objets."""
    ws = wb[tab_name]
    items = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        ref = cell_val(row[1])        # B: Référence
        origine = cell_val(row[2])    # C: I/C
        code_g = cell_val(row[3])     # D: Code granit
        granit = cell_val(row[4])     # E: Granit
        prix = cell_val(row[5])       # F: Prix HT

        if ref and prix is not None:
            items.append({
                "reference": str(ref),
                "origine": str(origine) if origine else "",
                "code_granit": int(code_g) if code_g else 0,
                "granit": str(granit) if granit else "",
                "prix_ht": clean_number(prix) if prix else 0,
            })
    return items


def read_accessoires(wb, tab_name):
    """Lit un onglet accessoires → liste d'objets."""
    ws = wb[tab_name]
    items = []
    photos_dir = os.path.join(PHOTOS_DIR, "accessoires")

    for row in ws.iter_rows(min_row=2, values_only=False):
        typ = cell_val(row[1])        # B: Type
        ref = cell_val(row[2])        # C: Référence
        origine = cell_val(row[3])    # D: I/C
        code_g = cell_val(row[4])     # E: Code granit
        granit = cell_val(row[5])     # F: Granit
        prix = cell_val(row[6])       # G: Prix HT

        if ref and prix is not None:
            entry = {
                "type": str(typ) if typ else "",
                "reference": str(ref),
                "origine": str(origine) if origine else "",
                "code_granit": int(code_g) if code_g else 0,
                "granit": str(granit) if granit else "",
                "prix_ht": clean_number(prix) if prix else 0,
            }
            photo = find_photo(photos_dir, normalize_ref(ref))
            if photo:
                entry["photo"] = photo
            items.append(entry)
    return items


def read_gravures(wb, tab_name):
    """Lit un onglet gravures → liste d'objets."""
    ws = wb[tab_name]
    items = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        ref = cell_val(row[1])        # B: Référence
        prix = cell_val(row[2])       # C: Prix caractère HT

        if ref and prix is not None:
            items.append({
                "reference": str(ref),
                "prix_caractere_ht": clean_number(prix) if prix else 0,
            })
    return items


def read_generic_product(wb, tab_name, product_type):
    """Lecture générique pour les onglets non reconnus (Litho, Urne, nouveaux types).
    Tente de détecter la structure à partir des headers.
    """
    ws = wb[tab_name]
    items = []

    # Lire les headers
    headers = []
    for cell in ws[1]:
        h = cell_val(cell)
        headers.append(str(h).lower().strip() if h else "")

    # Créer le sous-dossier photos si nécessaire
    type_lower = product_type.lower()
    if type_lower.endswith("e"):
        photos_subdir = type_lower + "s"
    else:
        photos_subdir = type_lower + "s"
    photos_dir = os.path.join(PHOTOS_DIR, photos_subdir)

    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell_val(c) for c in row]
        # Ignorer les lignes vides
        if not any(v is not None for v in vals):
            continue

        entry = {}
        for i, h in enumerate(headers):
            if i >= len(vals):
                break
            v = vals[i]
            if v is None:
                continue

            if "référence" in h or "reference" in h:
                entry["reference"] = str(v)
            elif "type" in h:
                entry["type"] = str(v)
            elif "ligne" in h:
                entry["ligne"] = str(v)
            elif h in ("i/c", "origine"):
                entry["origine"] = str(v)
            elif "code" in h and "granit" in h:
                entry["code_granit"] = int(v) if v else 0
            elif "code" in h:
                entry["code_granit"] = int(v) if v else 0
            elif "granit" in h:
                entry["granit"] = str(v)
            elif "caractère" in h or "caractere" in h:
                entry["prix_caractere_ht"] = clean_number(v) if v else 0
            elif "prix" in h:
                entry["prix_ht"] = clean_number(v) if v else 0

        if entry:
            # Photo
            ref = entry.get("reference", "")
            if ref:
                photo = find_photo(photos_dir, normalize_ref(ref))
                if photo:
                    entry["photo"] = photo
            items.append(entry)

    return items


# ============================================================
# ASSEMBLAGE
# ============================================================
def build_data():
    """Fonction principale : lit tout et assemble le data.json."""
    excel_path = find_excel()
    print(f"📂 Excel : {os.path.basename(excel_path)}")
    print(f"📁 Photos : {PHOTOS_DIR}")
    print()

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ---- Onglets structurels ----
    print("📋 Lecture des onglets structurels...")
    granits = read_granits(wb)
    poids = read_poids(wb)
    zones_transport = read_zones_transport(wb)
    tarifs_transport = read_tarifs_transport(wb)
    types_list, lignes_monument, lignes_accessoire, departements = read_listes(wb)

    granits_with_photo = sum(1 for g in granits if "photo" in g)
    print(f"  ✅ {len(granits)} granits ({granits_with_photo} avec photo)")
    print(f"  ✅ {len(poids)} poids")
    print(f"  ✅ {len(zones_transport)} départements → zones")
    print(f"  ✅ {len(tarifs_transport)} tarifs transport")
    print(f"  ✅ {len(departements)} départements")
    print(f"  ✅ Types : {types_list}")
    print(f"  ✅ Lignes monument : {lignes_monument}")
    print(f"  ✅ Lignes accessoire : {lignes_accessoire}")
    print()

    # ---- Onglets produits (auto-détection) ----
    product_tabs = detect_product_tabs(wb)
    print(f"🔍 Onglets produits détectés : {product_tabs}")
    print()

    data = {
        "granits": granits,
    }

    # Dictionnaires pour stocker les données par type
    for tab_name in product_tabs:
        product_type = extract_product_type(tab_name)
        print(f"📦 Lecture : {tab_name} (type: {product_type})...")

        if product_type == "Monument":
            items = read_monuments(wb, tab_name)
            key = "monuments"
            refs = set(i["reference"] for i in items)
            refs_with_photo = sum(1 for i in items if "photo" in i)
            unique_refs = len(refs)
            print(f"  ✅ {len(items)} lignes ({unique_refs} refs uniques, {refs_with_photo} lignes avec photo)")

        elif product_type == "Semelle":
            items = read_semelles(wb, tab_name)
            key = "semelles"
            print(f"  ✅ {len(items)} lignes")

        elif product_type == "Accessoire":
            items = read_accessoires(wb, tab_name)
            key = "accessoires"
            refs = set(i["reference"] for i in items)
            refs_with_photo = sum(1 for i in items if "photo" in i)
            print(f"  ✅ {len(items)} lignes ({len(refs)} refs uniques, {refs_with_photo} lignes avec photo)")

        elif product_type == "Gravure":
            items = read_gravures(wb, tab_name)
            key = "gravures"
            print(f"  ✅ {len(items)} lignes")

        else:
            # Type générique (Litho, Urne, ou nouveau type futur)
            items = read_generic_product(wb, tab_name, product_type)
            key = product_type.lower() + "s"
            print(f"  ✅ {len(items)} lignes (lecture générique)")

        data[key] = items

    # Assurer que les clés attendues existent même si l'onglet est vide
    for expected_key in ["monuments", "semelles", "accessoires", "gravures", "lithos", "urnes"]:
        if expected_key not in data:
            data[expected_key] = []
            print(f"  ⚠️  {expected_key} : onglet vide ou absent")

    # ---- Données structurelles ----
    data["poids"] = poids
    data["zones_transport"] = zones_transport
    data["tarifs_transport"] = tarifs_transport
    data["departements"] = departements
    data["types"] = types_list
    data["lignes_monument"] = lignes_monument
    data["lignes_accessoire"] = lignes_accessoire

    # ---- Écriture JSON ----
    print()
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    file_size = os.path.getsize(OUTPUT_FILE)
    print(f"✅ data.json généré : {OUTPUT_FILE}")
    print(f"   Taille : {file_size // 1024} Ko")

    # ---- Récap photos ----
    print()
    print("📸 État des dossiers photos :")
    photo_subdirs = ["monuments", "accessoires", "granits", "gravures", "lithos", "urnes"]
    # Ajouter les dossiers de nouveaux types détectés
    for tab_name in product_tabs:
        pt = extract_product_type(tab_name).lower() + "s"
        if pt not in photo_subdirs:
            photo_subdirs.append(pt)

    for subdir in photo_subdirs:
        path = os.path.join(PHOTOS_DIR, subdir)
        if os.path.isdir(path):
            photos = [f for f in os.listdir(path) if f.lower().endswith((".jpg", ".jpeg", ".png")) and not f.startswith(".")]
            if photos:
                print(f"  📁 {subdir}/ : {len(photos)} photo(s)")
            else:
                print(f"  📁 {subdir}/ : vide")
        else:
            print(f"  ⚠️  {subdir}/ : dossier manquant → création...")
            os.makedirs(path, exist_ok=True)

    print()
    print("🏁 Terminé.")
    return data


# ============================================================
# VÉRIFICATION (optionnel : --verify)
# ============================================================
def verify_against_html():
    """Compare le data.json généré avec les données du HTML existant."""
    html_path = None
    for f in os.listdir(SCRIPT_DIR):
        if f.endswith(".html") and "standalone" in f.lower():
            html_path = os.path.join(SCRIPT_DIR, f)
            break

    if not html_path:
        print("⚠️  Pas de fichier HTML standalone trouvé pour vérification.")
        return

    print(f"\n🔍 Vérification contre : {os.path.basename(html_path)}")

    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()

    with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
        new_data = json.load(f)

    # Extraire le JSON du HTML
    match = re.search(r"const DATA = (\{.*?\});\s*\n", html, re.DOTALL)
    if not match:
        print("❌ Impossible d'extraire const DATA du HTML.")
        return

    try:
        old_data = json.loads(match.group(1))
    except json.JSONDecodeError as e:
        print(f"❌ Erreur de parsing du DATA HTML : {e}")
        return

    # Comparer les clés
    old_keys = set(old_data.keys())
    new_keys = set(new_data.keys())

    if old_keys == new_keys:
        print("  ✅ Mêmes clés top-level")
    else:
        missing = old_keys - new_keys
        extra = new_keys - old_keys
        if missing:
            print(f"  ❌ Clés manquantes : {missing}")
        if extra:
            print(f"  ℹ️  Clés ajoutées : {extra}")

    # Comparer les comptages
    for key in sorted(old_keys & new_keys):
        old_val = old_data[key]
        new_val = new_data[key]

        if isinstance(old_val, list) and isinstance(new_val, list):
            if len(old_val) == len(new_val):
                print(f"  ✅ {key} : {len(old_val)} éléments")
            else:
                print(f"  ❌ {key} : HTML={len(old_val)}, JSON={len(new_val)}")
        elif isinstance(old_val, dict) and isinstance(new_val, dict):
            if len(old_val) == len(new_val):
                print(f"  ✅ {key} : {len(old_val)} entrées")
            else:
                print(f"  ❌ {key} : HTML={len(old_val)}, JSON={len(new_val)}")

    # Vérifier un monument au hasard
    if "monuments" in old_data and "monuments" in new_data and old_data["monuments"]:
        old_m = old_data["monuments"][0]
        new_m = new_data["monuments"][0]
        # Comparer les clés du premier monument (ignorer "photo" qui est nouveau)
        old_m_keys = set(old_m.keys())
        new_m_keys = set(k for k in new_m.keys() if k != "photo")
        if old_m_keys == new_m_keys:
            print(f"  ✅ Structure monument : clés identiques")
        else:
            print(f"  ❌ Structure monument diffère : HTML={old_m_keys}, JSON={new_m_keys}")

        # Comparer les valeurs du premier monument
        match_count = sum(1 for k in old_m_keys if old_m.get(k) == new_m.get(k))
        print(f"  ℹ️  Premier monument : {match_count}/{len(old_m_keys)} valeurs identiques")


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print("  PHG-France — Génération data.json")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    print()

    data = build_data()

    if "--verify" in sys.argv:
        verify_against_html()
